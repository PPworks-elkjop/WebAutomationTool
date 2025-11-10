"""
Create Excel template for ESL AP credentials
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path

def create_template():
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "AP Credentials"
    
    # Define headers
    headers = [
        "Retail Chain",
        "Store ID",
        "Store Alias",
        "AP ID",
        "IP Address",
        "Type",
        "Username Web UI",
        "Password Web UI",
        "Username SSH",
        "Password SSH",
        "SU Password",
        "Notes"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add example data
    example_data = [
        ["Elkjop", "S001", "Oslo Store", "AP-001", "192.168.1.100", "Indoor", "admin", "password123", "root", "sshpass123", "supass456", "Main store AP"],
        ["Elkjop", "S002", "Bergen Store", "AP-002", "192.168.1.101", "Outdoor", "admin", "password456", "root", "sshpass456", "supass789", "Warehouse AP"],
        ["Power", "S101", "Stockholm Store", "AP-101", "192.168.1.200", "Indoor", "admin", "pass001", "root", "sshroot", "supass001", ""],
    ]
    
    for row_num, row_data in enumerate(example_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws.cell(row=row_num, column=col_num).value = value
    
    # Set column widths
    column_widths = [15, 12, 20, 12, 16, 12, 18, 18, 15, 15, 15, 30]
    for col_num, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_num).column_letter].width = width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save template
    template_path = Path(__file__).parent / "AP_Credentials_Template.xlsx"
    wb.save(template_path)
    print(f"Template created: {template_path}")
    return template_path

if __name__ == "__main__":
    create_template()
