import tkinter as tk
from tkinter import ttk
from ping_manager import PingManager

class ConnectionStatusDialog:
    """Dialog showing real-time connection status for multiple APs."""
    
    def __init__(self, parent, ap_list, provisioning_callback=None, ssh_callback=None, 
                 close_browser_callback=None, ping_host_func=None):
        """Initialize the connection status dialog.
        
        Args:
            parent: Parent window
            ap_list: List of AP dictionaries to display
            provisioning_callback: Callback for Provisioning button
            ssh_callback: Callback for SSH button
            close_browser_callback: Callback for Close Browser button
            ping_host_func: Function to ping a host (ip_address, timeout) -> (success, response_time)
        """
        self.dialog = tk.Toplevel(parent)
        self.dialog.title("Connection Status")
        self.dialog.geometry("1400x750")
        self.dialog.configure(bg="#FFFFFF")
        
        # Make it non-modal (stay on top but don't block interaction)
        self.dialog.transient(parent)
        # Don't use grab_set() to allow interaction with other windows
        # Don't use topmost to avoid blocking other windows
        
        # Store AP list and callbacks
        self.ap_list = ap_list
        self.ap_rows = {}  # Map ap_id to tree item
        self.ap_selected = {}  # Track checkbox selections: ap_id -> bool
        self.provisioning_callback = provisioning_callback
        self.ssh_callback = ssh_callback
        self.close_browser_callback = close_browser_callback
        self.ping_host_func = ping_host_func
        self.reconnect_callback = None  # Will be set externally
        
        # Initialize ping manager
        self.ping_manager = PingManager(ping_host_func) if ping_host_func else None
        
        self._create_ui()
        self._populate_aps()
        
        # Center the dialog
        self.dialog.update_idletasks()
        x = (self.dialog.winfo_screenwidth() // 2) - (1400 // 2)
        y = (self.dialog.winfo_screenheight() // 2) - (750 // 2)
        self.dialog.geometry(f"1400x750+{x}+{y}")
    
    def _create_ui(self):
        """Create the UI elements."""
        # Header
        header_frame = tk.Frame(self.dialog, bg="#FFFFFF", pady=10)
        header_frame.pack(fill="x", padx=20)
        
        title_label = tk.Label(
            header_frame,
            text="Connection Progress",
            font=("Arial", 16, "bold"),
            bg="#FFFFFF"
        )
        title_label.pack()
        
        # Main content frame (holds button panel and tree)
        content_frame = tk.Frame(self.dialog, bg="#FFFFFF")
        content_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        # Left panel for action buttons
        button_panel = tk.Frame(content_frame, bg="#F8F9FA", relief="flat", bd=1)
        button_panel.pack(side="left", fill="y", padx=(0, 10))
        
        # Action buttons title
        tk.Label(
            button_panel,
            text="Actions",
            font=("Segoe UI", 11, "bold"),
            bg="#F8F9FA",
            fg="#333333"
        ).pack(pady=(10, 15))
        
        # Reconnect button (always visible)
        self.reconnect_btn = tk.Button(
            button_panel,
            text="Reconnect Selected",
            command=self._on_reconnect,
            font=("Segoe UI", 10),
            bg="#28A745",
            fg="white",
            activebackground="#218838",
            width=14,
            cursor="hand2",
            relief="flat",
            bd=0,
            padx=10,
            pady=8,
            state="disabled"
        )
        self.reconnect_btn.pack(pady=5, padx=10)
        
        # Provisioning button
        if self.provisioning_callback:
            self.provisioning_btn = tk.Button(
                button_panel,
                text="Provisioning",
                command=self._on_provisioning,
                font=("Segoe UI", 10),
                bg="#FFC107",
                fg="white",
                activebackground="#E0A800",
                width=14,
                cursor="hand2",
                relief="flat",
                bd=0,
                padx=10,
                pady=8,
                state="disabled"
            )
            self.provisioning_btn.pack(pady=5, padx=10)
        
        # SSH button
        if self.ssh_callback:
            self.ssh_btn = tk.Button(
                button_panel,
                text="SSH",
                command=self._on_ssh,
                font=("Segoe UI", 10),
                bg="#17A2B8",
                fg="white",
                activebackground="#138496",
                width=14,
                cursor="hand2",
                relief="flat",
                bd=0,
                padx=10,
                pady=8,
                state="disabled"
            )
            self.ssh_btn.pack(pady=5, padx=10)
        
        # Close Browser button
        if self.close_browser_callback:
            self.close_browser_btn = tk.Button(
                button_panel,
                text="Close Browser",
                command=self._on_close_browser,
                font=("Segoe UI", 10),
                bg="#6C757D",
                fg="white",
                activebackground="#5A6268",
                width=14,
                cursor="hand2",
                relief="flat",
                bd=0,
                padx=10,
                pady=8,
                state="disabled"
            )
            self.close_browser_btn.pack(pady=5, padx=10)
        
        # Separator
        ttk.Separator(button_panel, orient="horizontal").pack(fill="x", pady=15, padx=10)
        
        # Create treeview with scrollbar
        tree_frame = tk.Frame(content_frame, bg="#FFFFFF")
        tree_frame.pack(side="left", fill="both", expand=True)
        
        # Scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical")
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal")
        
        # Treeview with custom style for larger rows
        style = ttk.Style()
        style.configure("Custom.Treeview", rowheight=40, font=("Segoe UI", 10))
        style.configure("Custom.Treeview.Heading", font=("Segoe UI", 10, "bold"))
        
        # Create a custom font tag for the ping column with much larger size
        self.ping_font = ("Arial", 18, "bold")
        
        self.tree = ttk.Treeview(
            tree_frame,
            columns=("select", "store_id", "ap_id", "ip_address", "result", "status", "message", "ping_btn", "ping_result", "ping_count"),
            show="headings",
            yscrollcommand=vsb.set,
            xscrollcommand=hsb.set,
            height=20,
            style="Custom.Treeview"
        )
        
        vsb.config(command=self.tree.yview)
        hsb.config(command=self.tree.xview)
        
        # Configure columns
        self.tree.heading("select", text="☐")
        self.tree.heading("store_id", text="Store ID")
        self.tree.heading("ap_id", text="AP ID")
        self.tree.heading("ip_address", text="IP Address")
        self.tree.heading("result", text="SSH")  # SSH status indicator
        self.tree.heading("status", text="Status")
        self.tree.heading("message", text="Message")
        self.tree.heading("ping_btn", text="Ping")
        self.tree.heading("ping_result", text="Response")
        self.tree.heading("ping_count", text="#")
        
        self.tree.column("select", width=40, anchor="center")
        self.tree.column("store_id", width=80, anchor="center")
        self.tree.column("ap_id", width=120, anchor="center")
        self.tree.column("ip_address", width=120, anchor="center")
        self.tree.column("result", width=30, anchor="center")  # Narrow column for ✓/✗
        self.tree.column("status", width=100, anchor="center")
        self.tree.column("message", width=250, anchor="w")
        self.tree.column("ping_btn", width=60, anchor="center")
        self.tree.column("ping_result", width=100, anchor="center")
        self.tree.column("ping_count", width=40, anchor="center")
        
        # Configure tags for status colors
        self.tree.tag_configure("pending", background="#F8F9FA")
        self.tree.tag_configure("connecting", background="#FFF3CD")
        self.tree.tag_configure("connected", background="#D4EDDA")
        self.tree.tag_configure("failed", background="#F8D7DA")
        
        # Store full messages for tooltips
        self.full_messages = {}
        
        # Bind hover events for tooltips
        self.tree.bind("<Motion>", self._on_mouse_motion)
        self.tree.bind("<Leave>", self._on_mouse_leave)
        self.tree.bind("<Button-1>", self._on_tree_click)  # Handle ping button and checkbox clicks
        self.tooltip = None
        
        # Grid layout
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)
        
        # Add spacer to push export/hide buttons to bottom of left panel
        spacer = tk.Frame(button_panel, bg="#F8F9FA")
        spacer.pack(expand=True, fill="both")
        
        # Separator before bottom buttons
        ttk.Separator(button_panel, orient="horizontal").pack(fill="x", pady=10, padx=10)
        
        # Export to Excel button
        self.export_button = tk.Button(
            button_panel,
            text="📊 Export",
            command=self._on_export,
            font=("Segoe UI", 9),
            bg="#007BFF",
            fg="white",
            width=14,
            cursor="hand2",
            relief="flat",
            bd=0,
            padx=10,
            pady=6
        )
        self.export_button.pack(pady=5, padx=10)
        
        # Hide Window button
        self.close_button = tk.Button(
            button_panel,
            text="Hide Window",
            command=self._on_close,
            font=("Segoe UI", 9),
            bg="#6C757D",
            fg="white",
            width=14,
            cursor="hand2",
            relief="flat",
            bd=0,
            padx=10,
            pady=6
        )
        self.close_button.pack(pady=(5, 10), padx=10)
        
        # Summary label at bottom of window
        self.summary_label = tk.Label(
            self.dialog,
            text="Preparing to connect...",
            font=("Arial", 9),
            bg="#FFFFFF",
            fg="#666666"
        )
        self.summary_label.pack(side="bottom", pady=5)
    
    def _populate_aps(self):
        """Populate the treeview with APs."""
        for ap in self.ap_list:
            store_id = ap.get('store_id', 'Unknown')
            ap_id = ap.get('ap_id', 'Unknown')
            ip_address = ap.get('ip_address', 'Unknown')
            
            item = self.tree.insert(
                "",
                "end",
                values=("☐", store_id, ap_id, ip_address, "?", "Pending", "Waiting to connect...", "⏵", "", ""),
                tags=("pending",)
            )
            self.ap_rows[ap_id] = item
            self.ap_selected[ap_id] = False
    
    def update_status(self, ap_id, status, message="", result_indicator=None):
        """Update the status of an AP.
        
        Args:
            ap_id: AP identifier
            status: One of: "connecting", "connected", "failed"
            message: Status message to display
            result_indicator: Optional visual indicator ("✓", "✗", or None)
        """
        if ap_id not in self.ap_rows:
            return
        
        item = self.ap_rows[ap_id]
        current_values = self.tree.item(item, "values")
        
        # Map status to display text and tag
        status_map = {
            "connecting": ("Connecting...", "connecting"),
            "connected": ("✓ Connected", "connected"),
            "failed": ("✗ Failed", "failed")
        }
        
        status_text, tag = status_map.get(status, (status, "pending"))
        
        # Auto-detect result indicator from message if not provided
        # Logic: ✓ = SSH enabled, ✗ = SSH disabled, ? = not checked yet
        if result_indicator is None:
            msg_lower = message.lower()
            # Only update SSH indicator for SSH-related operations
            if "ssh" in msg_lower:
                # Check for explicit enabled/disabled states
                if "ssh is already enabled" in msg_lower or "ssh enabled successfully" in msg_lower:
                    result_indicator = "✓"
                elif "ssh has been enabled" in msg_lower or "ssh: enable: ssh" in msg_lower:
                    result_indicator = "✓"
                elif "ssh is currently enabled" in msg_lower and "disabled" not in msg_lower:
                    result_indicator = "✓"
                elif "ssh is currently disabled" in msg_lower or "ssh is already disabled" in msg_lower:
                    result_indicator = "✗"
                elif "ssh has been disabled" in msg_lower or "ssh: disable: ssh" in msg_lower:
                    result_indicator = "✗"
                else:
                    # SSH operation but unclear state (processing, error, etc)
                    result_indicator = None  # Keep current value
            else:
                # Non-SSH operation (connection, provisioning, etc) - don't change SSH status
                result_indicator = None  # Keep current value
        
        # Store full message for tooltip
        self.full_messages[item] = message
        
        # Truncate message for display (show first 50 chars)
        display_message = message if len(message) <= 50 else message[:47] + "..."
        
        # Preserve checkbox and ping columns
        checkbox = current_values[0] if len(current_values) > 0 else "☐"
        ping_btn = current_values[7] if len(current_values) > 7 else "⏵"
        ping_result = current_values[8] if len(current_values) > 8 else ""
        ping_count = current_values[9] if len(current_values) > 9 else ""
        
        # If result_indicator is None, preserve the current value (keep ? until SSH is checked)
        if result_indicator is None:
            result_indicator = current_values[4] if len(current_values) > 4 else "?"
        
        # Update the row
        self.tree.item(
            item,
            values=(checkbox, current_values[1], current_values[2], current_values[3], result_indicator, status_text, display_message, ping_btn, ping_result, ping_count),
            tags=(tag,)
        )
        
        # Scroll to show this item
        self.tree.see(item)
        
        # Update the dialog to show changes immediately
        self.dialog.update_idletasks()
    
    def update_summary(self, message):
        """Update the summary label.
        
        Args:
            message: Summary message to display
        """
        self.summary_label.config(text=message)
        self.dialog.update_idletasks()
    
    def enable_close(self):
        """Update the close button styling when operations are complete."""
        self.close_button.config(bg="#28A745")
        self.dialog.update_idletasks()
    
    def _on_export(self):
        """Export the status data to Excel."""
        try:
            from tkinter import filedialog, messagebox
            from datetime import datetime
            
            # Try to import openpyxl for Excel export
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font, Alignment, PatternFill
                use_excel = True
            except ImportError:
                use_excel = False
            
            if use_excel:
                # Export to Excel (.xlsx)
                default_filename = f"ap_status_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                filepath = filedialog.asksaveasfilename(
                    parent=self.dialog,
                    title="Export to Excel",
                    defaultextension=".xlsx",
                    filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                    initialfile=default_filename
                )
                
                if not filepath:
                    return
                
                # Create workbook and worksheet
                wb = Workbook()
                ws = wb.active
                ws.title = "AP Status"
                
                # Write header with styling
                headers = ["Store ID", "AP ID", "IP Address", "Enabled/Disabled", "Status", "Message"]
                ws.append(headers)
                
                # Style header row
                header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Write data rows
                for item in self.tree.get_children():
                    values = self.tree.item(item, "values")
                    
                    store_id = values[0] if len(values) > 0 else ""
                    ap_id = values[1] if len(values) > 1 else ""
                    ip_address = values[2] if len(values) > 2 else ""
                    result = values[3] if len(values) > 3 else ""
                    status = values[4] if len(values) > 4 else ""
                    full_msg = self.full_messages.get(item, values[5] if len(values) > 5 else "")
                    
                    ws.append([store_id, ap_id, ip_address, result, status, full_msg])
                
                # Auto-adjust column widths
                ws.column_dimensions['A'].width = 20  # Store ID
                ws.column_dimensions['B'].width = 15  # AP ID
                ws.column_dimensions['C'].width = 18  # IP Address
                ws.column_dimensions['D'].width = 12  # Enabled/Disabled
                ws.column_dimensions['E'].width = 15  # Status
                ws.column_dimensions['F'].width = 60  # Message
                
                # Save workbook
                wb.save(filepath)
                messagebox.showinfo("Export Successful", f"Data exported to:\n{filepath}", parent=self.dialog)
                
            else:
                # Fallback to CSV if openpyxl not available
                import csv
                default_filename = f"ap_status_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
                filepath = filedialog.asksaveasfilename(
                    parent=self.dialog,
                    title="Export to CSV",
                    defaultextension=".csv",
                    filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
                    initialfile=default_filename
                )
                
                if not filepath:
                    return
                
                with open(filepath, 'w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f, quoting=csv.QUOTE_ALL)
                    writer.writerow(["Store ID", "AP ID", "IP Address", "Enabled/Disabled", "Status", "Message"])
                    
                    for item in self.tree.get_children():
                        values = self.tree.item(item, "values")
                        store_id = values[0] if len(values) > 0 else ""
                        ap_id = values[1] if len(values) > 1 else ""
                        ip_address = values[2] if len(values) > 2 else ""
                        result = values[3] if len(values) > 3 else ""
                        status = values[4] if len(values) > 4 else ""
                        full_msg = self.full_messages.get(item, values[5] if len(values) > 5 else "")
                        clean_msg = full_msg.replace('\n', ' ').replace('\r', ' ').strip()
                        clean_msg = ' '.join(clean_msg.split())
                        writer.writerow([store_id, ap_id, ip_address, result, status, clean_msg])
                
                messagebox.showinfo("Export Successful", f"Data exported to:\n{filepath}\n\nNote: Install 'openpyxl' for native Excel format export.", parent=self.dialog)
            
        except Exception as e:
            from tkinter import messagebox
            messagebox.showerror("Export Failed", f"Failed to export data:\n{str(e)}", parent=self.dialog)
    
    def _on_close(self):
        """Hide the dialog (don't destroy it so we can show it again)."""
        self.dialog.withdraw()
    
    def show_window(self):
        """Show the dialog window."""
        self.dialog.deiconify()
        self.dialog.lift()
    
    def _on_mouse_motion(self, event):
        """Show tooltip on mouse hover."""
        item = self.tree.identify_row(event.y)
        if item and item in self.full_messages:
            # Get the full message
            full_message = self.full_messages[item]
            
            # Only show tooltip if message is truncated
            if len(full_message) > 50:
                # Destroy existing tooltip
                self._destroy_tooltip()
                
                # Create new tooltip
                x = event.x_root + 10
                y = event.y_root + 10
                
                self.tooltip = tk.Toplevel(self.dialog)
                self.tooltip.wm_overrideredirect(True)
                self.tooltip.wm_geometry(f"+{x}+{y}")
                # Tooltip should be above other windows
                self.tooltip.attributes('-topmost', True)
                
                label = tk.Label(
                    self.tooltip,
                    text=full_message,
                    background="#FFFFE0",
                    relief="solid",
                    borderwidth=1,
                    font=("Arial", 9),
                    padx=5,
                    pady=3,
                    wraplength=400
                )
                label.pack()
        else:
            self._destroy_tooltip()
    
    def _on_mouse_leave(self, event):
        """Hide tooltip when mouse leaves."""
        self._destroy_tooltip()
    
    def _destroy_tooltip(self):
        """Destroy the tooltip if it exists."""
        if self.tooltip:
            try:
                self.tooltip.destroy()
            except:
                pass
            self.tooltip = None
    
    def _on_provisioning(self):
        """Handle Provisioning button click."""
        if self.provisioning_callback:
            self.provisioning_callback()
    
    def _on_ssh(self):
        """Handle SSH button click."""
        if self.ssh_callback:
            self.ssh_callback()
    
    def _on_close_browser(self):
        """Handle Close Browser button click."""
        if self.close_browser_callback:
            self.close_browser_callback()
    
    def enable_action_buttons(self):
        """Enable the action buttons (Provisioning, SSH, Close Browser)."""
        if hasattr(self, 'provisioning_btn'):
            self.provisioning_btn.config(state="normal")
        if hasattr(self, 'ssh_btn'):
            self.ssh_btn.config(state="normal")
        if hasattr(self, 'close_browser_btn'):
            self.close_browser_btn.config(state="normal")
    
    def disable_action_buttons(self):
        """Disable the action buttons (Provisioning, SSH, Close Browser)."""
        if hasattr(self, 'provisioning_btn'):
            self.provisioning_btn.config(state="disabled")
        if hasattr(self, 'ssh_btn'):
            self.ssh_btn.config(state="disabled")
        if hasattr(self, 'close_browser_btn'):
            self.close_browser_btn.config(state="disabled")
    
    def _on_tree_click(self, event):
        """Handle clicks on the tree - specifically on checkboxes and ping buttons."""
        region = self.tree.identify_region(event.x, event.y)
        if region not in ["cell", "tree"]:
            return
        
        column = self.tree.identify_column(event.x)
        item = self.tree.identify_row(event.y)
        
        if not item or not column:
            return
        
        values = self.tree.item(item, "values")
        
        # Checkbox is column #1
        if column == "#1":
            if len(values) >= 3:
                ap_id = values[2]  # ap_id is in column 2 (was 1)
                self._toggle_checkbox(ap_id, item)
        
        # Ping button is column index 7 (0-based), which is #8 in Tkinter (was #7)
        elif column == "#8":
            if len(values) >= 4:
                ap_id = values[2]  # ap_id is in column 2 (was 1)
                ip_address = values[3]  # ip_address is in column 3 (was 2)
                
                # Toggle ping state
                if self.ping_manager and self.ping_manager.is_pinging(ap_id):
                    # Stop ping
                    self._stop_ping(ap_id, item)
                else:
                    # Start ping
                    self._start_ping(ap_id, ip_address, item)
    
    def _toggle_checkbox(self, ap_id, item):
        """Toggle checkbox selection for an AP."""
        # Toggle selection state
        self.ap_selected[ap_id] = not self.ap_selected.get(ap_id, False)
        
        # Update checkbox display
        current_values = list(self.tree.item(item, "values"))
        current_values[0] = "☑" if self.ap_selected[ap_id] else "☐"
        self.tree.item(item, values=current_values)
        
        # Update reconnect button state
        self._update_reconnect_button()
    
    def _update_reconnect_button(self):
        """Enable/disable reconnect button based on selections."""
        has_selection = any(self.ap_selected.values())
        if has_selection:
            self.reconnect_btn.config(state="normal")
        else:
            self.reconnect_btn.config(state="disabled")
    
    def _on_reconnect(self):
        """Handle Reconnect button click - reconnect to selected APs."""
        # Get list of selected AP IDs
        selected_ap_ids = [ap_id for ap_id, selected in self.ap_selected.items() if selected]
        
        if not selected_ap_ids:
            return
        
        # Get full AP data for selected APs
        selected_aps = [ap for ap in self.ap_list if ap.get('ap_id') in selected_ap_ids]
        
        # Call reconnect callback if set
        if self.reconnect_callback:
            self.reconnect_callback(selected_aps)
        
        # Reset status for selected APs to "Pending"
        for ap_id in selected_ap_ids:
            if ap_id in self.ap_rows:
                item = self.ap_rows[ap_id]
                current_values = list(self.tree.item(item, "values"))
                # Preserve checkbox, keep other initial values
                checkbox = current_values[0]
                current_values = [checkbox, current_values[1], current_values[2], current_values[3], "?", "Pending", "Reconnecting...", "⏵", "", ""]
                self.tree.item(item, values=current_values, tags=("pending",))
    
    def _start_ping(self, ap_id, ip_address, item):
        """Start continuous ping for an AP."""
        if not self.ping_manager:
            print("No ping manager available")
            return
        
        # Update button to show "⏸"
        current_values = list(self.tree.item(item, "values"))
        current_values[7] = "⏸"
        current_values[8] = "Pinging..."
        current_values[9] = "0"
        self.tree.item(item, values=current_values)
        
        # Create update callback
        def update_callback(result_text, count):
            def update_ui():
                if ap_id not in self.ap_rows:
                    return
                
                try:
                    current_values = list(self.tree.item(item, "values"))
                    current_values[8] = result_text
                    current_values[9] = str(count)
                    self.tree.item(item, values=current_values)
                    
                    # Update background color based on result
                    if "Timeout" in result_text or "Error" in result_text:
                        self.tree.tag_configure(f"ping_fail_{ap_id}", background="#F8D7DA")
                        current_tags = list(self.tree.item(item, "tags"))
                        current_tags = [t for t in current_tags if not t.startswith("ping_")]
                        current_tags.append(f"ping_fail_{ap_id}")
                        self.tree.item(item, tags=current_tags)
                    else:
                        self.tree.tag_configure(f"ping_ok_{ap_id}", background="#D4EDDA")
                        current_tags = list(self.tree.item(item, "tags"))
                        current_tags = [t for t in current_tags if not t.startswith("ping_")]
                        current_tags.append(f"ping_ok_{ap_id}")
                        self.tree.item(item, tags=current_tags)
                except:
                    pass
            
            try:
                self.dialog.after(0, update_ui)
            except:
                pass
        
        # Start ping using manager
        self.ping_manager.start_ping(ap_id, ip_address, update_callback)
    
    def _stop_ping(self, ap_id, item):
        """Stop continuous ping for an AP."""
        if self.ping_manager:
            self.ping_manager.stop_ping(ap_id)
        
        # Update button to show "⏵"
        current_values = list(self.tree.item(item, "values"))
        current_values[7] = "⏵"
        self.tree.item(item, values=current_values)
    
    def destroy(self):
        """Destroy the dialog."""
        try:
            # Stop all ping threads using manager
            if self.ping_manager:
                self.ping_manager.stop_all()
            
            self._destroy_tooltip()
            self.dialog.destroy()
        except:
            pass
