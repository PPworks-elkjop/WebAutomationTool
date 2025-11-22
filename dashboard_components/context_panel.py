"""
Context Panel - Upper Right
Shows contextual lists based on active AP (Jira tickets, Vusion data)
"""

import tkinter as tk
from tkinter import ttk, messagebox, scrolledtext


class ContextPanel:
    """Upper right panel - Contextual lists (Jira, Vusion) for active AP."""
    
    def __init__(self, parent, db, current_user=None, on_selection=None, log_callback=None):
        self.parent = parent
        self.db = db
        self.current_user = current_user
        self.on_selection = on_selection
        self.log_callback = log_callback
        
        self.active_ap = None
        self.active_ap_data = None
        
        self._create_ui()
    
    def _create_ui(self):
        """Create context panel UI."""
        # Header
        header = tk.Frame(self.parent, bg="#3D6B9E", height=40)
        header.pack(fill=tk.X, side=tk.TOP)
        header.pack_propagate(False)
        
        self.header_label = tk.Label(header, text="Contextual Data", font=('Segoe UI', 12, 'bold'),
                                     bg="#3D6B9E", fg="white")
        self.header_label.pack(side=tk.LEFT, padx=15, pady=8)
        
        # Import CustomNotebook
        import sys
        import os
        sys.path.insert(0, os.path.dirname(os.path.dirname(__file__)))
        from custom_notebook import CustomNotebook
        
        # Notebook for different context types with custom styling
        self.notebook = CustomNotebook(self.parent, tab_font=('Segoe UI', 11), tab_height=36)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Notes tab (first)
        self.notes_frame = tk.Frame(self.notebook.content_area, bg="#FFFFFF")
        self._populate_notes_tab(self.notes_frame)
        self.notebook.add(self.notes_frame, text="Notes")
        
        # Jira tab
        self.jira_frame = tk.Frame(self.notebook.content_area, bg="#FFFFFF")
        self._populate_jira_tab(self.jira_frame)
        self.notebook.add(self.jira_frame, text="Jira Tickets")
        
        # Vusion tab
        self.vusion_frame = tk.Frame(self.notebook.content_area, bg="#FFFFFF")
        self._populate_vusion_tab(self.vusion_frame)
        self.notebook.add(self.vusion_frame, text="Vusion Integration")
        
        # Show placeholder
        self._show_placeholder()
    
    def _populate_jira_tab(self, frame):
        """Populate Jira tickets tab."""
        # Placeholder container
        self.jira_placeholder_frame = tk.Frame(frame, bg="#FFFFFF")
        self.jira_placeholder_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(self.jira_placeholder_frame, text="Select an item to view data",
                font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(expand=True)
        
        # Content frame (initially hidden)
        self.jira_content_frame = tk.Frame(frame, bg="#FFFFFF")
        
        # Filters panel
        filters_frame = tk.Frame(self.jira_content_frame, bg="#F8F9FA", padx=10, pady=10)
        filters_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Top row with buttons on the right
        top_row = tk.Frame(filters_frame, bg="#F8F9FA")
        top_row.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(top_row, text="Filters", font=('Segoe UI', 11, 'bold'),
                bg="#F8F9FA", fg="#212529").pack(side=tk.LEFT)
        
        # Buttons on the right
        tk.Button(top_row, text="🔄 Refresh", command=self._refresh_jira,
                 bg="#6C757D", fg="white", font=('Segoe UI', 9),
                 padx=12, pady=6, relief=tk.FLAT, cursor="hand2",
                 borderwidth=0).pack(side=tk.RIGHT, padx=(5, 0))
        
        tk.Button(top_row, text="Apply Filters", command=self._apply_jira_filters,
                 bg="#3D6B9E", fg="white", font=('Segoe UI', 9, 'bold'),
                 padx=15, pady=6, relief=tk.FLAT, cursor="hand2",
                 borderwidth=0).pack(side=tk.RIGHT)
        
        # Combined row: Ticket ID and Date Range
        search_row = tk.Frame(filters_frame, bg="#F8F9FA")
        search_row.pack(fill=tk.X, pady=(0, 8))
        
        # Ticket ID section
        tk.Label(search_row, text="Ticket ID:", font=('Segoe UI', 9, 'bold'),
                bg="#F8F9FA", fg="#495057").pack(side=tk.LEFT, padx=(0, 5))
        
        ticket_entry_container = tk.Frame(search_row, bg="#FFFFFF", relief=tk.SOLID, borderwidth=1)
        ticket_entry_container.pack(side=tk.LEFT, padx=(0, 15))
        
        self.jira_ticket_id = tk.Entry(ticket_entry_container, font=('Segoe UI', 9), width=15,
                                       bg="#FFFFFF", fg="#212529", relief=tk.FLAT, bd=0)
        self.jira_ticket_id.pack(side=tk.LEFT, padx=5, pady=3)
        self.jira_ticket_id.insert(0, "e.g., FIXIT-1192609")
        self.jira_ticket_id.config(fg="#6C757D")
        
        # Add placeholder behavior
        def on_ticket_focus_in(event):
            if self.jira_ticket_id.get() == "e.g., FIXIT-1192609":
                self.jira_ticket_id.delete(0, tk.END)
                self.jira_ticket_id.config(fg="#212529")
        
        def on_ticket_focus_out(event):
            if not self.jira_ticket_id.get():
                self.jira_ticket_id.insert(0, "e.g., FIXIT-1192609")
                self.jira_ticket_id.config(fg="#6C757D")
        
        self.jira_ticket_id.bind("<FocusIn>", on_ticket_focus_in)
        self.jira_ticket_id.bind("<FocusOut>", on_ticket_focus_out)
        self.jira_ticket_id.bind("<Return>", lambda e: self._apply_jira_filters())
        
        # Date range section (on same row)
        tk.Label(search_row, text="Date Range:", font=('Segoe UI', 9, 'bold'),
                bg="#F8F9FA", fg="#495057").pack(side=tk.LEFT, padx=(0, 5))
        
        tk.Label(search_row, text="From:", font=('Segoe UI', 9),
                bg="#F8F9FA", fg="#495057").pack(side=tk.LEFT, padx=(0, 5))
        
        # Custom styled date from entry with calendar button
        date_from_container = tk.Frame(search_row, bg="#FFFFFF", relief=tk.SOLID, borderwidth=1)
        date_from_container.pack(side=tk.LEFT, padx=(0, 5))
        
        self.jira_date_from = tk.Entry(date_from_container, font=('Segoe UI', 9), width=10,
                                       bg="#FFFFFF", fg="#212529", relief=tk.FLAT, bd=0)
        self.jira_date_from.pack(side=tk.LEFT, padx=5, pady=3)
        
        cal_from_btn = tk.Label(date_from_container, text="📅", font=('Segoe UI', 10),
                               bg="#FFFFFF", fg="#495057", cursor="hand2")
        cal_from_btn.pack(side=tk.LEFT, padx=(0, 5))
        cal_from_btn.bind("<Button-1>", lambda e: self._show_date_picker(self.jira_date_from))
        
        tk.Label(search_row, text="To:", font=('Segoe UI', 9),
                bg="#F8F9FA", fg="#495057").pack(side=tk.LEFT, padx=(5, 5))
        
        # Custom styled date to entry with calendar button
        date_to_container = tk.Frame(search_row, bg="#FFFFFF", relief=tk.SOLID, borderwidth=1)
        date_to_container.pack(side=tk.LEFT)
        
        self.jira_date_to = tk.Entry(date_to_container, font=('Segoe UI', 9), width=10,
                                     bg="#FFFFFF", fg="#212529", relief=tk.FLAT, bd=0)
        self.jira_date_to.pack(side=tk.LEFT, padx=5, pady=3)
        
        cal_to_btn = tk.Label(date_to_container, text="📅", font=('Segoe UI', 10),
                             bg="#FFFFFF", fg="#495057", cursor="hand2")
        cal_to_btn.pack(side=tk.LEFT, padx=(0, 5))
        cal_to_btn.bind("<Button-1>", lambda e: self._show_date_picker(self.jira_date_to))
        
        # Project filter
        project_row = tk.Frame(filters_frame, bg="#F8F9FA")
        project_row.pack(fill=tk.X, pady=(0, 8))
        
        tk.Label(project_row, text="Projects:", font=('Segoe UI', 9, 'bold'),
                bg="#F8F9FA", fg="#495057", width=12, anchor="w").pack(side=tk.LEFT)
        
        self.jira_projects = {}  # Will store {project_key: BooleanVar}
        self.project_checkboxes_frame = tk.Frame(project_row, bg="#F8F9FA")
        self.project_checkboxes_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Status filter
        status_row = tk.Frame(filters_frame, bg="#F8F9FA")
        status_row.pack(fill=tk.X, pady=(0, 5))
        
        tk.Label(status_row, text="Status:", font=('Segoe UI', 9, 'bold'),
                bg="#F8F9FA", fg="#495057", width=12, anchor="w").pack(side=tk.LEFT)
        
        self.jira_statuses = {}
        self.status_checkboxes_frame = tk.Frame(status_row, bg="#F8F9FA")
        self.status_checkboxes_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        # Common statuses with custom checkboxes - start with common ones
        # Note: This will be updated dynamically based on found tickets
        for status in ['Open', 'In Progress', 'Waiting for Support', 'Resolved', 'Closed']:
            var = tk.BooleanVar(value=True)
            self.jira_statuses[status] = var
            self._create_custom_checkbox(self.status_checkboxes_frame, status, var)
        
        # Scrollable list for tickets with custom layout
        list_container = tk.Frame(self.jira_content_frame, bg="#FFFFFF")
        list_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        canvas = tk.Canvas(list_container, bg="#FFFFFF", highlightthickness=0)
        scrollbar = tk.Scrollbar(list_container, orient="vertical", command=canvas.yview)
        self.jira_tickets_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        self.jira_tickets_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=self.jira_tickets_frame, anchor="nw", width=canvas.winfo_width())
        canvas.configure(yscrollcommand=scrollbar.set)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Store canvas for width updates
        self.jira_canvas = canvas
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(canvas.find_withtag("all")[0], width=e.width) if canvas.find_withtag("all") else None)
        
        # Store ticket data
        self.jira_tickets = []
    
    def _populate_notes_tab(self, frame):
        """Populate notes tab for active AP."""
        # Main container
        self.notes_main_frame = tk.Frame(frame, bg="#FFFFFF", padx=20, pady=15)
        self.notes_main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Header with title and Add Note button
        self.notes_header_frame = tk.Frame(self.notes_main_frame, bg="#FFFFFF")
        self.notes_header_frame.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(self.notes_header_frame, text="Notes", font=('Segoe UI', 11, 'bold'),
                bg="#FFFFFF", fg="#333333").pack(side=tk.LEFT)
        
        self.add_note_btn = tk.Button(self.notes_header_frame, text="Add Note", 
                                      command=self._trigger_add_note,
                                      bg="#28A745", fg="white", cursor="hand2", padx=15, pady=6,
                                      font=('Segoe UI', 9, 'bold'), relief=tk.FLAT, bd=0,
                                      activebackground="#218838", state=tk.DISABLED)
        self.add_note_btn.pack(side=tk.RIGHT)
        
        # Separator line
        tk.Frame(self.notes_main_frame, bg="#CCCCCC", height=1).pack(fill=tk.X, pady=(0, 10))
        
        # Notes list with scrollbar
        self.notes_list_frame = tk.Frame(self.notes_main_frame, bg="#FFFFFF")
        self.notes_list_frame.pack(fill=tk.BOTH, expand=True)
        
        self.notes_canvas = tk.Canvas(self.notes_list_frame, bg="#FFFFFF", highlightthickness=0)
        self.notes_scrollbar = tk.Scrollbar(self.notes_list_frame, orient="vertical", 
                                           command=self.notes_canvas.yview)
        self.notes_container = tk.Frame(self.notes_canvas, bg="#FFFFFF")
        
        def _update_scroll_region(event):
            self.notes_canvas.configure(scrollregion=self.notes_canvas.bbox("all"))
            canvas_width = self.notes_canvas.winfo_width()
            self.notes_canvas.itemconfig(self.notes_window, width=canvas_width)
        
        self.notes_container.bind("<Configure>", _update_scroll_region)
        self.notes_canvas.bind("<Configure>", _update_scroll_region)
        
        self.notes_window = self.notes_canvas.create_window((0, 0), window=self.notes_container, anchor="nw")
        self.notes_canvas.configure(yscrollcommand=self.notes_scrollbar.set)
        
        self.notes_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.notes_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            self.notes_canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        def _bind_mousewheel(event):
            self.notes_canvas.bind("<MouseWheel>", _on_mousewheel)
        
        def _unbind_mousewheel(event):
            self.notes_canvas.unbind("<MouseWheel>")
        
        self.notes_canvas.bind("<Enter>", _bind_mousewheel)
        self.notes_canvas.bind("<Leave>", _unbind_mousewheel)
        
        # Show placeholder initially
        tk.Label(self.notes_container, text="Select an item to view data",
                font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(pady=20)
    
    def _populate_vusion_tab(self, frame):
        """Populate Vusion integration tab."""
        # Placeholder container
        self.vusion_placeholder_frame = tk.Frame(frame, bg="#FFFFFF")
        self.vusion_placeholder_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(self.vusion_placeholder_frame, text="Select an item to view data",
                font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(expand=True)
        
        # Content frame (initially hidden)
        self.vusion_content_frame = tk.Frame(frame, bg="#FFFFFF")
        
        # Filters panel
        filters_frame = tk.Frame(self.vusion_content_frame, bg="#F8F9FA", padx=10, pady=10)
        filters_frame.pack(fill=tk.X, padx=5, pady=5)
        
        # Top row with title and buttons
        top_row = tk.Frame(filters_frame, bg="#F8F9FA")
        top_row.pack(fill=tk.X, pady=(0, 10))
        
        tk.Label(top_row, text="Event Filters", font=('Segoe UI', 11, 'bold'),
                bg="#F8F9FA", fg="#212529").pack(side=tk.LEFT)
        
        tk.Button(top_row, text="🔄 Refresh", command=self._refresh_vusion_events,
                 bg="#6C757D", fg="white", font=('Segoe UI', 9),
                 padx=12, pady=6, relief=tk.FLAT, cursor="hand2",
                 borderwidth=0).pack(side=tk.RIGHT)
        
        tk.Button(top_row, text="📊 Export to Excel", command=self._export_vusion_to_excel,
                 bg="#28A745", fg="white", font=('Segoe UI', 9),
                 padx=12, pady=6, relief=tk.FLAT, cursor="hand2",
                 borderwidth=0).pack(side=tk.RIGHT, padx=(0, 5))
        
        tk.Button(top_row, text="📋 Copy Table", command=self._copy_vusion_table,
                 bg="#007BFF", fg="white", font=('Segoe UI', 9),
                 padx=12, pady=6, relief=tk.FLAT, cursor="hand2",
                 borderwidth=0).pack(side=tk.RIGHT, padx=(0, 5))
        
        # Event type filters
        filter_row = tk.Frame(filters_frame, bg="#F8F9FA")
        filter_row.pack(fill=tk.X)
        
        tk.Label(filter_row, text="Event Type:", font=('Segoe UI', 9, 'bold'),
                bg="#F8F9FA", fg="#495057").pack(side=tk.LEFT, padx=(0, 10))
        
        # Checkboxes for event types
        self.vusion_event_types = {}
        self.vusion_event_checkboxes_frame = tk.Frame(filter_row, bg="#F8F9FA")
        self.vusion_event_checkboxes_frame.pack(side=tk.LEFT, fill=tk.X, expand=True)
        
        for event_type in ['Transmitter Connectivity', 'Transmitter Configuration']:
            var = tk.BooleanVar(value=True)
            self.vusion_event_types[event_type] = var
            # Add command to refresh display when filter changes
            var.trace_add('write', lambda *args: self._apply_vusion_filters())
            self._create_custom_checkbox(self.vusion_event_checkboxes_frame, event_type, var)
        
        # Events table
        table_container = tk.Frame(self.vusion_content_frame, bg="#FFFFFF")
        table_container.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        # Create canvas for scrolling (vertical only, table adapts to width)
        canvas = tk.Canvas(table_container, bg="#FFFFFF", highlightthickness=0)
        scrollbar_y = tk.Scrollbar(table_container, orient="vertical", command=canvas.yview)
        
        self.vusion_events_frame = tk.Frame(canvas, bg="#FFFFFF")
        
        # Configure canvas to resize frame to fit width
        def _configure_canvas(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
            # Make the frame fill the canvas width
            canvas_width = event.width
            canvas.itemconfig(self.vusion_canvas_window, width=canvas_width)
        
        self.vusion_events_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        self.vusion_canvas_window = canvas.create_window((0, 0), window=self.vusion_events_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar_y.set)
        canvas.bind("<Configure>", _configure_canvas)
        
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_y.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Store canvas reference
        self.vusion_canvas = canvas
        
        # Mouse wheel scrolling
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), "units")
        
        canvas.bind("<Enter>", lambda e: canvas.bind("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind("<MouseWheel>"))
        
        # Store events data
        self.vusion_events = []
    
    def set_active_ap(self, ap_id, ap_data):
        """Update context panel for a new active AP."""
        self.active_ap = ap_id
        self.active_ap_data = ap_data
        
        self.header_label.config(text=f"Context: AP {ap_id}")
        
        # Load notes for this AP (fast, no API calls)
        self._load_notes()
        
        # Load Jira tickets in background to avoid UI freezing
        self._load_jira_tickets_background()
        
        # Load Vusion data for this AP (fast, no API calls)
        self._load_vusion_data()
        
        self._log(f"Context updated for AP {ap_id}")
    
    def _load_notes(self):
        """Load notes for active AP."""
        # Clear existing notes
        for widget in self.notes_container.winfo_children():
            widget.destroy()
        
        if not self.active_ap:
            tk.Label(self.notes_container, text="Select an item to view data",
                    font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(pady=20)
            self.add_note_btn.config(state=tk.DISABLED)
            return
        
        # Enable Add Note button
        self.add_note_btn.config(state=tk.NORMAL)
        
        try:
            # Get notes from database
            notes = self.db.get_support_notes(self.active_ap)
            
            if notes:
                for note in notes:
                    self._create_note_item(self.notes_container, note)
                self._log(f"Loaded {len(notes)} notes for AP {self.active_ap}")
            else:
                tk.Label(self.notes_container, text="No notes found for this AP",
                        font=('Segoe UI', 9, 'italic'), bg="#FFFFFF", fg="#888888").pack(pady=20)
        except Exception as e:
            tk.Label(self.notes_container, text=f'Error loading notes: {str(e)}',
                    font=('Segoe UI', 9), bg="#FFFFFF", fg="#DC3545").pack(pady=20)
            self._log(f"Error loading notes: {str(e)}", "error")
    
    def _load_jira_tickets_background(self):
        """Load Jira tickets asynchronously to avoid UI freezing."""
        # Show loading message immediately
        for widget in self.jira_tickets_frame.winfo_children():
            widget.destroy()
        
        loading_label = tk.Label(self.jira_tickets_frame, 
                                text="🔄 Loading Jira tickets...",
                                font=('Segoe UI', 10), bg="#FFFFFF", fg="#6C757D")
        loading_label.pack(pady=20)
        
        # Use after() to run loading on main thread but yield to UI
        self.parent.after(50, self._load_jira_tickets)
    
    def _load_jira_tickets(self):
        """Load Jira tickets related to active AP with filters."""
        if not self.active_ap:
            # Show placeholder
            self.jira_content_frame.pack_forget()
            self.jira_placeholder_frame.pack(fill=tk.BOTH, expand=True)
            return
        
        # Setup UI
        self.jira_placeholder_frame.pack_forget()
        self.jira_content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Clear existing tickets
        for widget in self.jira_tickets_frame.winfo_children():
            widget.destroy()
        self.jira_tickets = []
        
        try:
            from jira_api import JiraAPI
            from credentials_manager import CredentialsManager
            
            # Initialize Jira API
            credentials_manager = CredentialsManager(self.db)
            jira_api = JiraAPI(credentials_manager)
            
            if not jira_api.is_configured():
                tk.Label(self.jira_tickets_frame, text="Jira not configured. Go to Admin Settings.",
                        font=('Segoe UI', 10), bg="#FFFFFF", fg="#DC3545").pack(pady=20)
                self._log("Jira not configured", "warning")
                return
            
            # Build JQL query - check if user specified a ticket ID first
            ticket_id = self.jira_ticket_id.get().strip()
            if ticket_id and ticket_id != "e.g., FIXIT-1192609" and '-' in ticket_id:
                # User specified a ticket ID - search for that specifically
                # Don't apply date filters when searching by ticket ID
                jql = f'key = "{ticket_id.upper()}"'
                self._log(f"Searching for specific ticket: {ticket_id.upper()}")
            else:
                # Regular AP ID search - use multiple fields for better coverage
                # Add wildcard to numeric AP IDs (like Jira does) to find variations
                search_term = f"{self.active_ap}*" if self.active_ap.isdigit() else self.active_ap
                jql = f'(text ~ "{search_term}" OR summary ~ "{search_term}" OR description ~ "{search_term}" OR comment ~ "{search_term}")'
                
                # Get date filters (only for AP ID searches)
                date_from = self.jira_date_from.get().strip()
                date_to = self.jira_date_to.get().strip()
                
                # Add date range to JQL
                if date_from:
                    jql += f' AND created >= "{date_from}"'
                if date_to:
                    jql += f' AND created <= "{date_to} 23:59"'
            
            self._log(f"Jira JQL query: {jql}")
            success, result, message = jira_api.search_issues(jql, max_results=200)
            self._log(f"Jira search result - success: {success}, message: {message}")
            
            # If searching by ticket ID and not found, provide detailed feedback
            if ticket_id and ticket_id != "e.g., FIXIT-1192609" and '-' in ticket_id:
                if not success or not result.get('issues'):
                    self._log(f"Ticket {ticket_id.upper()} not found via JQL search")
                    # Try direct API call
                    try:
                        success_direct, issue_data, msg = jira_api.get_issue(ticket_id.upper())
                        if success_direct and issue_data:
                            self._log(f"Ticket {ticket_id.upper()} found via direct API but not via search")
                            # Wrap in expected format
                            result = {'issues': [issue_data], 'total': 1}
                            success = True
                        else:
                            self._log(f"Ticket {ticket_id.upper()} not found via direct API either: {msg}")
                    except Exception as e:
                        self._log(f"Error trying direct API: {str(e)}")
            
            # Also try without date filter to see if we get any results
            if ticket_id and ticket_id != "e.g., FIXIT-1192609" and '-' in ticket_id:
                simple_jql = f'key = "{ticket_id.upper()}"'
            else:
                search_term = f"{self.active_ap}*" if self.active_ap.isdigit() else self.active_ap
                simple_jql = f'(text ~ "{search_term}" OR summary ~ "{search_term}" OR description ~ "{search_term}" OR comment ~ "{search_term}")'
            success2, result2, message2 = jira_api.search_issues(simple_jql, max_results=200)
            self._log(f"Jira search WITHOUT date filter - success: {success2}, message: {message2}")
            if success2 and isinstance(result2, dict):
                issues2 = result2.get('issues', [])
                self._log(f"Found {len(issues2)} issues without date filter")
            
            if success and isinstance(result, dict):
                issues = result.get('issues', [])
                self._log(f"Found {len(issues)} Jira issues (with date filter)")
                
                # Update project and status checkboxes based on ALL found issues
                self._update_project_filters(issues)
                self._update_status_filters(issues)
                
                # Apply client-side filters
                filtered_issues = []
                enabled_projects = [proj for proj, var in self.jira_projects.items() if var.get()]
                enabled_statuses = [status for status, var in self.jira_statuses.items() if var.get()]
                
                for issue in issues:
                    fields = issue.get('fields', {})
                    
                    # Check project filter
                    project = fields.get('project', {})
                    project_key = project.get('key', '') if isinstance(project, dict) else ''
                    if enabled_projects and project_key not in enabled_projects:
                        continue
                    
                    # Check status filter
                    status = fields.get('status', {})
                    status_name = status.get('name', '') if isinstance(status, dict) else ''
                    if enabled_statuses and status_name not in enabled_statuses:
                        continue
                    
                    filtered_issues.append(issue)
                
                # Display tickets
                if filtered_issues:
                    for issue in filtered_issues:
                        self._create_jira_ticket_card(issue)
                        self.jira_tickets.append(issue)
                    self._log(f"Loaded {len(filtered_issues)} Jira tickets (filtered from {len(issues)} total)")
                else:
                    # Show appropriate message based on search type
                    if ticket_id and ticket_id != "e.g., FIXIT-1192609":
                        msg = f"Ticket '{ticket_id.upper()}' not found.\n\nPlease verify:\n• Ticket exists\n• You have access to it\n• Ticket ID is correct"
                    else:
                        msg = f"No Jira tickets found matching filters for '{self.active_ap}'"
                    
                    tk.Label(self.jira_tickets_frame, text=msg,
                            font=('Segoe UI', 10), bg="#FFFFFF", fg="#6C757D",
                            justify=tk.LEFT).pack(pady=20)
                    self._log("No Jira tickets found")
            else:
                tk.Label(self.jira_tickets_frame, text=f"Error: {message}",
                        font=('Segoe UI', 10), bg="#FFFFFF", fg="#DC3545").pack(pady=20)
                self._log(f"Jira search error: {message}", "error")
                
        except Exception as e:
            tk.Label(self.jira_tickets_frame, text=f"Jira error: {str(e)}",
                    font=('Segoe UI', 10), bg="#FFFFFF", fg="#DC3545").pack(pady=20)
            self._log(f"Jira error: {str(e)}", "error")
    
    def _create_custom_checkbox(self, parent, text, variable):
        """Create a custom styled checkbox."""
        container = tk.Frame(parent, bg="#F8F9FA")
        container.pack(side=tk.LEFT, padx=(0, 15))
        
        checkbox = tk.Frame(container, bg="#FFFFFF", relief=tk.SOLID, borderwidth=1,
                           width=16, height=16, cursor="hand2")
        checkbox.pack(side=tk.LEFT, padx=(0, 5))
        checkbox.pack_propagate(False)
        
        check_label = tk.Label(checkbox, text="", bg="#FFFFFF", fg="#3D6B9E", font=('Segoe UI', 10, 'bold'))
        check_label.pack(expand=True)
        
        text_label = tk.Label(container, text=text, font=('Segoe UI', 9),
                             bg="#F8F9FA", fg="#495057", cursor="hand2")
        text_label.pack(side=tk.LEFT)
        
        def toggle():
            variable.set(not variable.get())
            update_display()
        
        def update_display():
            if variable.get():
                check_label.config(text="✓")
                checkbox.config(bg="#3D6B9E", borderwidth=1)
                check_label.config(bg="#3D6B9E", fg="#FFFFFF")
            else:
                check_label.config(text="")
                checkbox.config(bg="#FFFFFF", borderwidth=1)
                check_label.config(bg="#FFFFFF")
        
        checkbox.bind("<Button-1>", lambda e: toggle())
        check_label.bind("<Button-1>", lambda e: toggle())
        text_label.bind("<Button-1>", lambda e: toggle())
        
        update_display()
    
    def _show_date_picker(self, entry_widget):
        """Show a simple date picker dialog."""
        from datetime import datetime, timedelta
        import calendar
        
        picker = tk.Toplevel(self.parent)
        picker.title("Select Date")
        picker.geometry("280x320")
        picker.transient(self.parent)
        
        # Wait for window to be ready before grabbing
        picker.update_idletasks()
        try:
            picker.grab_set()
        except:
            pass  # Ignore grab errors
        
        # Try to parse current date from entry
        try:
            current_date = datetime.strptime(entry_widget.get(), '%Y-%m-%d')
        except:
            current_date = datetime.now()
        
        selected_date = [current_date]
        
        def update_calendar():
            for widget in cal_frame.winfo_children():
                widget.destroy()
            
            year = selected_date[0].year
            month = selected_date[0].month
            
            # Month/Year header
            header = tk.Frame(cal_frame, bg="#3D6B9E")
            header.pack(fill=tk.X)
            
            tk.Button(header, text="◀", command=prev_month, bg="#3D6B9E", fg="white",
                     font=('Segoe UI', 10), relief=tk.FLAT, cursor="hand2", padx=10).pack(side=tk.LEFT)
            
            tk.Label(header, text=f"{calendar.month_name[month]} {year}",
                    font=('Segoe UI', 11, 'bold'), bg="#3D6B9E", fg="white").pack(side=tk.LEFT, expand=True)
            
            tk.Button(header, text="▶", command=next_month, bg="#3D6B9E", fg="white",
                     font=('Segoe UI', 10), relief=tk.FLAT, cursor="hand2", padx=10).pack(side=tk.RIGHT)
            
            # Day headers
            days_header = tk.Frame(cal_frame, bg="#F8F9FA")
            days_header.pack(fill=tk.X)
            for day in ['Mo', 'Tu', 'We', 'Th', 'Fr', 'Sa', 'Su']:
                tk.Label(days_header, text=day, font=('Segoe UI', 8, 'bold'),
                        bg="#F8F9FA", fg="#495057", width=4).pack(side=tk.LEFT, expand=True)
            
            # Calendar days
            cal = calendar.monthcalendar(year, month)
            for week in cal:
                week_frame = tk.Frame(cal_frame, bg="#FFFFFF")
                week_frame.pack(fill=tk.X)
                for day in week:
                    if day == 0:
                        tk.Label(week_frame, text="", width=4, bg="#FFFFFF").pack(side=tk.LEFT, expand=True)
                    else:
                        day_btn = tk.Label(week_frame, text=str(day), font=('Segoe UI', 9),
                                         bg="#FFFFFF", fg="#212529", width=4, cursor="hand2",
                                         relief=tk.FLAT, padx=2, pady=4)
                        day_btn.pack(side=tk.LEFT, expand=True)
                        
                        # Highlight today
                        if (day == selected_date[0].day and month == selected_date[0].month 
                            and year == selected_date[0].year):
                            day_btn.config(bg="#3D6B9E", fg="white", font=('Segoe UI', 9, 'bold'))
                        
                        day_btn.bind("<Button-1>", lambda e, d=day: select_day(d))
                        day_btn.bind("<Enter>", lambda e, btn=day_btn: btn.config(bg="#E9ECEF") 
                                    if btn.cget("bg") == "#FFFFFF" else None)
                        day_btn.bind("<Leave>", lambda e, btn=day_btn: btn.config(bg="#FFFFFF") 
                                    if btn.cget("bg") == "#E9ECEF" else None)
        
        def prev_month():
            year = selected_date[0].year
            month = selected_date[0].month - 1
            if month == 0:
                month = 12
                year -= 1
            selected_date[0] = selected_date[0].replace(year=year, month=month, day=1)
            update_calendar()
        
        def next_month():
            year = selected_date[0].year
            month = selected_date[0].month + 1
            if month == 13:
                month = 1
                year += 1
            selected_date[0] = selected_date[0].replace(year=year, month=month, day=1)
            update_calendar()
        
        def select_day(day):
            selected_date[0] = selected_date[0].replace(day=day)
            entry_widget.delete(0, tk.END)
            entry_widget.insert(0, selected_date[0].strftime('%Y-%m-%d'))
            picker.destroy()
        
        cal_frame = tk.Frame(picker, bg="#FFFFFF")
        cal_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Buttons
        btn_frame = tk.Frame(picker, bg="#F8F9FA")
        btn_frame.pack(fill=tk.X, padx=10, pady=(0, 10))
        
        tk.Button(btn_frame, text="Today", command=lambda: select_day(datetime.now().day),
                 bg="#6C757D", fg="white", font=('Segoe UI', 9),
                 padx=15, pady=5, relief=tk.FLAT, cursor="hand2").pack(side=tk.LEFT)
        
        tk.Button(btn_frame, text="Cancel", command=picker.destroy,
                 bg="#DC3545", fg="white", font=('Segoe UI', 9),
                 padx=15, pady=5, relief=tk.FLAT, cursor="hand2").pack(side=tk.RIGHT)
        
        update_calendar()
    
    def _apply_jira_filters(self):
        """Apply filters and reload Jira tickets."""
        self._log("Applying Jira filters...")
        self._load_jira_tickets_background()
    
    def _update_project_filters(self, issues):
        """Update project filter checkboxes based on found issues."""
        # Extract unique projects from issues
        projects = set()
        for issue in issues:
            fields = issue.get('fields', {})
            project = fields.get('project', {})
            if isinstance(project, dict):
                project_key = project.get('key')
                if project_key:
                    projects.add(project_key)
        
        # Add new project checkboxes if needed
        for project_key in projects:
            if project_key not in self.jira_projects:
                var = tk.BooleanVar(value=True)
                self.jira_projects[project_key] = var
                self._create_custom_checkbox(self.project_checkboxes_frame, project_key, var)
    
    def _update_status_filters(self, issues):
        """Update status filter checkboxes based on found issues."""
        # Extract unique statuses from issues
        statuses = set()
        for issue in issues:
            fields = issue.get('fields', {})
            status = fields.get('status', {})
            if isinstance(status, dict):
                status_name = status.get('name')
                if status_name:
                    statuses.add(status_name)
        
        # Add new status checkboxes if needed (check exact key to avoid duplicates)
        for status_name in statuses:
            if status_name not in self.jira_statuses:
                self._log(f"Adding new status filter: {status_name}")
                var = tk.BooleanVar(value=True)
                self.jira_statuses[status_name] = var
                self._create_custom_checkbox(self.status_checkboxes_frame, status_name, var)
            else:
                self._log(f"Status filter already exists: {status_name}")
    
    def _create_jira_ticket_card(self, issue):
        """Create a card for a Jira ticket with status badge."""
        key = issue.get('key', 'Unknown')
        fields = issue.get('fields', {})
        summary = fields.get('summary', 'No summary')
        status = fields.get('status', {})
        status_name = status.get('name', 'Unknown') if isinstance(status, dict) else 'Unknown'
        
        # Card frame
        card = tk.Frame(self.jira_tickets_frame, bg="#F8F9FA", relief=tk.SOLID, borderwidth=1)
        card.pack(fill=tk.X, padx=5, pady=3)
        
        # Make card clickable
        def on_click(e=None):
            self._log(f"Selected Jira ticket: {key}")
            if self.on_selection:
                self.on_selection("jira", issue)
        
        card.bind("<Button-1>", on_click)
        
        # Add hover effect
        def on_enter(e):
            card.config(bg="#E9ECEF")
            content_frame.config(bg="#E9ECEF")
        
        def on_leave(e):
            card.config(bg="#F8F9FA")
            content_frame.config(bg="#F8F9FA")
        
        card.bind("<Enter>", on_enter)
        card.bind("<Leave>", on_leave)
        
        # Content
        content_frame = tk.Frame(card, bg="#F8F9FA")
        content_frame.pack(fill=tk.X, padx=10, pady=8)
        content_frame.bind("<Button-1>", on_click)
        
        # Top row: Key and Status
        top_row = tk.Frame(content_frame, bg="#F8F9FA")
        top_row.pack(fill=tk.X, pady=(0, 5))
        top_row.bind("<Button-1>", on_click)
        
        key_label = tk.Label(top_row, text=key, font=('Segoe UI', 10, 'bold'),
                            bg="#F8F9FA", fg="#0066CC", cursor="hand2")
        key_label.pack(side=tk.LEFT)
        key_label.bind("<Button-1>", on_click)
        
        # Status badge
        status_color = "#28A745" if status_name.lower() in ['done', 'resolved', 'closed'] else "#007BFF"
        status_frame = tk.Frame(top_row, bg=status_color, padx=6, pady=2)
        status_frame.pack(side=tk.RIGHT)
        status_frame.bind("<Button-1>", on_click)
        
        status_label = tk.Label(status_frame, text=status_name, font=('Segoe UI', 7, 'bold'),
                               bg=status_color, fg="white")
        status_label.pack()
        status_label.bind("<Button-1>", on_click)
        
        # Summary
        summary_label = tk.Label(content_frame, text=summary, font=('Segoe UI', 9),
                                bg="#F8F9FA", fg="#495057", anchor="w", justify=tk.LEFT,
                                wraplength=350, cursor="hand2")
        summary_label.pack(fill=tk.X)
        summary_label.bind("<Button-1>", on_click)
    
    def _load_vusion_data(self):
        """Load Vusion events for active AP."""
        if not self.active_ap_data:
            # Show placeholder, hide content
            self.vusion_content_frame.pack_forget()
            self.vusion_placeholder_frame.pack(fill=tk.BOTH, expand=True)
            return
        
        # Hide placeholder, show content
        self.vusion_placeholder_frame.pack_forget()
        self.vusion_content_frame.pack(fill=tk.BOTH, expand=True)
        
        # Load events in background thread
        self._load_vusion_events_background()
    
    def _load_vusion_events_background(self):
        """Load Vusion events asynchronously."""
        # Show loading message
        for widget in self.vusion_events_frame.winfo_children():
            widget.destroy()
        
        loading_label = tk.Label(self.vusion_events_frame, 
                                text="🔄 Loading Vusion events...",
                                font=('Segoe UI', 10), bg="#FFFFFF", fg="#6C757D")
        loading_label.pack(pady=20)
        
        # Load in background thread
        import threading
        thread = threading.Thread(target=self._load_vusion_events_thread, daemon=True)
        thread.start()
    
    def _load_vusion_events_thread(self):
        """Background thread to load Vusion events."""
        try:
            from vusion_api_helper import VusionAPIHelper
            from vusion_api_config import VusionAPIConfig
            
            store_id = self.active_ap_data.get('store_id', '')
            if not store_id:
                self.parent.after(0, lambda: self._show_vusion_error("No store ID available"))
                return
            
            # Parse country from store_id
            country = self._get_country_from_store_id(store_id)
            if not country:
                self.parent.after(0, lambda: self._show_vusion_error("Could not determine country from store ID"))
                return
            
            # Check if API key is configured
            config = VusionAPIConfig()
            api_key = config.get_api_key(country, 'vusion_pro')
            if not api_key:
                self.parent.after(0, lambda: self._show_vusion_error("Vusion API not configured for this country"))
                return
            
            # Get events from Vusion API
            helper = VusionAPIHelper()
            success, result = helper.get_events(country, store_id, search=self.active_ap, page_size=100)
            
            if success and isinstance(result, dict):
                # The API returns events in 'values' array, not 'content'
                events = result.get('values', [])
                self.parent.after(0, lambda: self._display_vusion_events(events))
            else:
                error_msg = result if isinstance(result, str) else "Failed to load events"
                self.parent.after(0, lambda msg=error_msg: self._show_vusion_error(msg))
        
        except Exception as e:
            self.parent.after(0, lambda: self._show_vusion_error(f"Error: {str(e)}"))
    
    def _get_country_from_store_id(self, store_id):
        """Parse country code from store_id."""
        if not store_id:
            return None
        
        store_lower = store_id.lower()
        
        # Check for lab environment first
        if 'elkjop_se_lab' in store_lower:
            return 'LAB'
        
        # Parse country from store_id pattern
        if '_no' in store_lower:
            return 'NO'
        elif '_se' in store_lower:
            return 'SE'
        elif '_fi' in store_lower:
            return 'FI'
        elif '_dk' in store_lower:
            return 'DK'
        elif '_is' in store_lower:
            return 'IS'
        
        return None
    
    def _display_vusion_events(self, events):
        """Display Vusion events in the table."""
        # Clear existing content
        for widget in self.vusion_events_frame.winfo_children():
            widget.destroy()
        
        self.vusion_events = events
        
        if not events:
            tk.Label(self.vusion_events_frame, text="No events found for this AP",
                    font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(pady=20)
            return
        
        # Apply filters
        filtered_events = self._filter_vusion_events(events)
        
        if not filtered_events:
            tk.Label(self.vusion_events_frame, text="No events match the selected filters",
                    font=('Segoe UI', 10, 'italic'), bg="#FFFFFF", fg="#6C757D").pack(pady=20)
            return
        
        # Create table header
        header_frame = tk.Frame(self.vusion_events_frame, bg="#E9ECEF", relief=tk.SOLID, borderwidth=1)
        header_frame.pack(fill=tk.X, padx=0, pady=(0, 0))
        
        # Configure grid columns
        header_frame.grid_columnconfigure(0, weight=2, minsize=150)  # Date
        header_frame.grid_columnconfigure(1, weight=2, minsize=200)  # Event Type
        header_frame.grid_columnconfigure(2, weight=1, minsize=100)  # Status
        header_frame.grid_columnconfigure(3, weight=4, minsize=300)  # Message
        
        # Header labels
        headers = [
            ('Date', 0),
            ('Event Type', 1),
            ('Status', 2),
            ('Message', 3)
        ]
        
        for header_text, col in headers:
            tk.Label(header_frame, text=header_text, font=('Segoe UI', 9, 'bold'),
                    bg="#E9ECEF", fg="#212529", anchor="w", padx=10, pady=8).grid(
                        row=0, column=col, sticky="ew")
        
        # Create event rows
        for idx, event in enumerate(filtered_events):
            self._create_vusion_event_row(self.vusion_events_frame, event, idx)
        
        self._log(f"Loaded {len(filtered_events)} Vusion events (filtered from {len(events)})")
    
    def _filter_vusion_events(self, events):
        """Filter events based on selected event types."""
        filtered = []
        
        # Get selected event types (UI names)
        selected_types = [et for et, var in self.vusion_event_types.items() if var.get()]
        
        # Map UI names to API event types
        type_mapping = {
            'Transmitter Connectivity': 'TRANSMITTER_CONNECTIVITY',
            'Transmitter Configuration': 'TRANSMITTER_CONFIGURATION'
        }
        
        # Convert to API event type names
        selected_api_types = [type_mapping.get(st, st) for st in selected_types]
        
        for event in events:
            event_type = event.get('eventType', '')
            
            # Check if event type matches any selected filter
            if event_type in selected_api_types:
                filtered.append(event)
        
        return filtered
    
    def _create_vusion_event_row(self, parent, event, index):
        """Create a row for a Vusion event."""
        # Alternate row colors
        bg_color = "#FFFFFF" if index % 2 == 0 else "#F8F9FA"
        
        row_frame = tk.Frame(parent, bg=bg_color, relief=tk.SOLID, borderwidth=1)
        row_frame.pack(fill=tk.X, padx=0, pady=0)
        
        # Configure grid columns (same as header)
        row_frame.grid_columnconfigure(0, weight=2, minsize=150)
        row_frame.grid_columnconfigure(1, weight=2, minsize=200)
        row_frame.grid_columnconfigure(2, weight=1, minsize=100)
        row_frame.grid_columnconfigure(3, weight=4, minsize=300)
        
        # Extract event data
        # API uses 'creationDate' not 'eventDate'
        date_str = event.get('creationDate', 'N/A')
        if date_str != 'N/A' and 'T' in date_str:
            # Format: 2025-11-19T17:42:01.207Z -> 2025-11-19 17:42:01
            date_str = date_str.replace('T', ' ').split('.')[0]
        
        # Convert API event type to user-friendly format
        event_type_raw = event.get('eventType', 'N/A')
        # Convert TRANSMITTER_CONNECTIVITY to "Transmitter Connectivity"
        event_type = event_type_raw.replace('_', ' ').title() if event_type_raw != 'N/A' else 'N/A'
        
        # Status might not be present in all events, derive from modifications if needed
        status = event.get('status', '')
        if not status and 'modifications' in event:
            # For connectivity events, show the new connectivity status
            new_val = event.get('modifications', {}).get('newValue', {})
            if 'connectivity.status' in new_val:
                status = new_val['connectivity.status']
        if not status:
            status = 'N/A'
        
        message = event.get('message', 'N/A')
        
        # Create labels with wrapping for long text
        date_label = tk.Label(row_frame, text=date_str, font=('Segoe UI', 9),
                             bg=bg_color, fg="#212529", anchor="nw", justify=tk.LEFT,
                             wraplength=140, padx=10, pady=8)
        date_label.grid(row=0, column=0, sticky="new")
        
        type_label = tk.Label(row_frame, text=event_type, font=('Segoe UI', 9),
                             bg=bg_color, fg="#212529", anchor="nw", justify=tk.LEFT,
                             wraplength=190, padx=10, pady=8)
        type_label.grid(row=0, column=1, sticky="new")
        
        # Status with color coding
        status_upper = status.upper()
        if status_upper == "ONLINE" or status_upper == "SUCCESS":
            status_color = "#28A745"  # Green
        elif status_upper == "OFFLINE" or status_upper == "ERROR":
            status_color = "#DC3545"  # Red
        else:
            status_color = "#6C757D"  # Gray
        
        status_label = tk.Label(row_frame, text=status, font=('Segoe UI', 9, 'bold'),
                               bg=bg_color, fg=status_color, anchor="nw", justify=tk.LEFT,
                               wraplength=90, padx=10, pady=8)
        status_label.grid(row=0, column=2, sticky="new")
        
        message_label = tk.Label(row_frame, text=message, font=('Segoe UI', 9),
                                bg=bg_color, fg="#212529", anchor="nw", justify=tk.LEFT,
                                wraplength=290, padx=10, pady=8, cursor="hand2")
        message_label.grid(row=0, column=3, sticky="new")
        
        # Make labels copyable
        for label, value in [(date_label, date_str), (type_label, event_type), 
                            (status_label, status), (message_label, message)]:
            self._make_vusion_label_copyable(label, value, bg_color)
    
    def _make_vusion_label_copyable(self, label, value, bg_color):
        """Make a label copyable with click and hover effects."""
        def copy_value(e=None):
            if value and value != 'N/A':
                self.parent.clipboard_clear()
                self.parent.clipboard_append(value)
                original_fg = label['fg']
                label.config(fg="#28A745")
                self.parent.after(500, lambda: label.config(fg=original_fg))
        
        label.bind("<Button-1>", copy_value)
        
        def on_enter(e):
            if value != 'N/A':
                label.config(bg="#E9ECEF")
        
        def on_leave(e):
            label.config(bg=bg_color)
        
        label.bind("<Enter>", on_enter)
        label.bind("<Leave>", on_leave)
    
    def _show_vusion_error(self, error_message):
        """Show error message in Vusion tab."""
        for widget in self.vusion_events_frame.winfo_children():
            widget.destroy()
        
        tk.Label(self.vusion_events_frame, text=f"⚠ {error_message}",
                font=('Segoe UI', 10), bg="#FFFFFF", fg="#DC3545").pack(pady=20)
    
    def _refresh_vusion_events(self):
        """Refresh Vusion events list."""
        self._log("Refreshing Vusion events...")
        self._load_vusion_events_background()
    
    def _apply_vusion_filters(self):
        """Apply filters to Vusion events without reloading from API."""
        if hasattr(self, 'vusion_events') and self.vusion_events:
            self._display_vusion_events(self.vusion_events)
    
    def _copy_vusion_table(self):
        """Copy entire Vusion events table to clipboard in tab-separated format."""
        if not hasattr(self, 'vusion_events') or not self.vusion_events:
            self._log("No events to copy", "warning")
            return
        
        # Apply filters to get currently displayed events
        filtered_events = self._filter_vusion_events(self.vusion_events)
        
        if not filtered_events:
            self._log("No events match current filters", "warning")
            return
        
        # Build tab-separated table
        lines = []
        
        # Header row
        lines.append("Date\tEvent Type\tStatus\tMessage")
        
        # Data rows
        for event in filtered_events:
            # Format date
            date_str = event.get('creationDate', 'N/A')
            if date_str != 'N/A' and 'T' in date_str:
                date_str = date_str.replace('T', ' ').split('.')[0]
            
            # Format event type
            event_type_raw = event.get('eventType', 'N/A')
            event_type = event_type_raw.replace('_', ' ').title() if event_type_raw != 'N/A' else 'N/A'
            
            # Get status
            status = event.get('status', '')
            if not status and 'modifications' in event:
                new_val = event.get('modifications', {}).get('newValue', {})
                if 'connectivity.status' in new_val:
                    status = new_val['connectivity.status']
            if not status:
                status = 'N/A'
            
            # Get message
            message = event.get('message', 'N/A')
            
            # Add row (escape tabs and newlines in message)
            message_clean = message.replace('\t', '  ').replace('\n', ' ').replace('\r', '')
            lines.append(f"{date_str}\t{event_type}\t{status}\t{message_clean}")
        
        # Copy to clipboard
        table_text = '\n'.join(lines)
        self.parent.clipboard_clear()
        self.parent.clipboard_append(table_text)
        
        self._log(f"Copied {len(filtered_events)} events to clipboard", "success")
    
    def _export_vusion_to_excel(self):
        """Export Vusion events to Excel file with AP ID as sheet name."""
        if not hasattr(self, 'vusion_events') or not self.vusion_events:
            self._log("No events to export", "warning")
            messagebox.showwarning("No Data", "No Vusion events to export.")
            return
        
        # Apply filters to get currently displayed events
        filtered_events = self._filter_vusion_events(self.vusion_events)
        
        if not filtered_events:
            self._log("No events match current filters", "warning")
            messagebox.showwarning("No Data", "No events match current filters.")
            return
        
        # Try to import openpyxl
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            self._log("openpyxl not installed", "error")
            messagebox.showerror("Missing Dependency", 
                               "openpyxl is not installed. Install it using:\npip install openpyxl")
            return
        
        # Get AP ID for sheet name
        ap_id = "Unknown"
        if self.active_ap_data:
            ap_id = str(self.active_ap_data.get('ap_id', 'Unknown'))
        
        # Ask user for save location
        from tkinter import filedialog
        default_filename = f"Vusion_Events_AP_{ap_id}.xlsx"
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename,
            title="Save Vusion Events"
        )
        
        if not filepath:
            return  # User cancelled
        
        try:
            # Create workbook and worksheet with AP ID as sheet name
            wb = Workbook()
            ws = wb.active
            # Sanitize sheet name (Excel limits: 31 chars, no special chars)
            sheet_name = f"AP {ap_id}"[:31]
            ws.title = sheet_name
            
            # Header styling
            header_fill = PatternFill(start_color="3D6B9E", end_color="3D6B9E", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_alignment = Alignment(horizontal="left", vertical="center")
            
            # Write headers
            headers = ["Date", "Event Type", "Status", "Message"]
            for col_idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
            
            # Set column widths
            ws.column_dimensions['A'].width = 20  # Date
            ws.column_dimensions['B'].width = 30  # Event Type
            ws.column_dimensions['C'].width = 15  # Status
            ws.column_dimensions['D'].width = 50  # Message
            
            # Write data rows
            for row_idx, event in enumerate(filtered_events, start=2):
                # Format date
                date_str = event.get('creationDate', 'N/A')
                if date_str != 'N/A' and 'T' in date_str:
                    date_str = date_str.replace('T', ' ').split('.')[0]
                
                # Format event type
                event_type_raw = event.get('eventType', 'N/A')
                event_type = event_type_raw.replace('_', ' ').title() if event_type_raw != 'N/A' else 'N/A'
                
                # Get status
                status = event.get('status', '')
                if not status and 'modifications' in event:
                    new_val = event.get('modifications', {}).get('newValue', {})
                    if 'connectivity.status' in new_val:
                        status = new_val['connectivity.status']
                if not status:
                    status = 'N/A'
                
                # Get message
                message = event.get('message', 'N/A')
                
                # Write cells
                ws.cell(row=row_idx, column=1, value=date_str)
                ws.cell(row=row_idx, column=2, value=event_type)
                
                # Status cell with color
                status_cell = ws.cell(row=row_idx, column=3, value=status)
                status_upper = status.upper()
                if status_upper in ['ONLINE', 'SUCCESS']:
                    status_cell.font = Font(bold=True, color="28A745")
                elif status_upper in ['OFFLINE', 'ERROR']:
                    status_cell.font = Font(bold=True, color="DC3545")
                
                ws.cell(row=row_idx, column=4, value=message)
                
                # Set alignment for all cells in row
                for col_idx in range(1, 5):
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(
                        horizontal="left", vertical="top", wrap_text=True
                    )
            
            # Save workbook
            wb.save(filepath)
            self._log(f"Exported {len(filtered_events)} events to {filepath}", "success")
            messagebox.showinfo("Export Successful", 
                              f"Exported {len(filtered_events)} events to:\n{filepath}")
        
        except Exception as e:
            self._log(f"Failed to export to Excel: {str(e)}", "error")
            messagebox.showerror("Export Failed", f"Failed to export to Excel:\n{str(e)}")
    
    def _refresh_jira(self):
        """Refresh Jira ticket list."""
        self._log("Refreshing Jira tickets...")
        self._load_jira_tickets_background()
    
    def _show_placeholder(self):
        """Show placeholder when no AP is active."""
        # Show all placeholders, hide all content frames
        self.jira_content_frame.pack_forget()
        self.jira_placeholder_frame.pack(fill=tk.BOTH, expand=True)
        
        self.vusion_content_frame.pack_forget()
        self.vusion_placeholder_frame.pack(fill=tk.BOTH, expand=True)
        
        # Notes placeholder is already showing from initialization
    
    def _create_note_item(self, parent, note):
        """Create a note item in 2-row format (date/user/replies, then headline)."""
        note_frame = tk.Frame(parent, bg="#FFFFFF", cursor="hand2")
        note_frame.pack(fill=tk.X, pady=3, padx=0)
        
        # Row 1: Date/Time, User (full name), and Reply count
        row1_frame = tk.Frame(note_frame, bg="#FFFFFF")
        row1_frame.pack(fill=tk.X, padx=0, pady=(5, 0))
        
        # Get user full name
        user_info = self.db.get_user(note['user'])
        display_name = user_info.get('full_name', note['user']) if user_info else note['user']
        
        tk.Label(row1_frame, text=f"{note['created_at']} - {display_name}",
                font=('Segoe UI', 8), fg="#888888", bg="#FFFFFF", anchor="w").pack(side=tk.LEFT)
        
        # Get reply count
        reply_count = self.db.get_note_reply_count(note['id'])
        if True:  # Always show reply count (even if 0)
            reply_frame = tk.Frame(row1_frame, bg="#FFFFFF")
            reply_frame.pack(side=tk.RIGHT, padx=5)
            
            # Forum icon (using emoji)
            icon_label = tk.Label(reply_frame, text="💬", font=('Segoe UI', 14),
                                 fg="#007BFF", bg="#FFFFFF", cursor="hand2")
            icon_label.pack(side=tk.LEFT, padx=(0, 3))
            
            # Reply count
            count_label = tk.Label(reply_frame, text=str(reply_count),
                                  font=('Segoe UI', 8), fg="#007BFF", bg="#FFFFFF",
                                  cursor="hand2")
            count_label.pack(side=tk.LEFT)
            
            # Tooltip
            def show_tooltip(event, widget):
                tooltip = tk.Toplevel()
                tooltip.wm_overrideredirect(True)
                tooltip.wm_geometry(f"+{event.x_root+10}+{event.y_root+10}")
                label = tk.Label(tooltip, text="Replies", background="#FFFFE0", relief=tk.SOLID,
                               borderwidth=1, font=('Segoe UI', 9), padx=4, pady=2)
                label.pack()
                widget.tooltip_window = tooltip
            
            def hide_tooltip(event, widget):
                if hasattr(widget, 'tooltip_window'):
                    widget.tooltip_window.destroy()
                    del widget.tooltip_window
            
            for widget in [icon_label, count_label]:
                widget.bind("<Enter>", lambda e, w=widget: show_tooltip(e, w))
                widget.bind("<Leave>", lambda e, w=widget: hide_tooltip(e, w))
        
        # Row 2: Headline
        row2 = tk.Label(note_frame, text=note['headline'],
                       font=('Segoe UI', 9, 'bold'), bg="#FFFFFF", anchor="w", fg="#333333")
        row2.pack(fill=tk.X, padx=0, pady=(0, 5))
        
        # Separator line
        tk.Frame(note_frame, bg="#E0E0E0", height=1).pack(fill=tk.X, pady=(5, 0))
        
        # Bind click events to open note details
        for widget in [note_frame, row1_frame, row2]:
            widget.bind("<Button-1>", lambda e, n=note: self._open_note_window(n))
            for child in widget.winfo_children():
                child.bind("<Button-1>", lambda e, n=note: self._open_note_window(n))
    
    def _trigger_add_note(self):
        """Trigger add note form in content panel."""
        if not self.active_ap:
            messagebox.showwarning("No AP Selected", "Please select an AP first.")
            return
        
        # Call content panel via selection callback
        if self.on_selection:
            self.on_selection("add_note", {"ap_id": self.active_ap, "ap_data": self.active_ap_data})
    
    def _open_write_note_dialog(self):
        """Open dialog to write a new note."""
        if not self.active_ap:
            messagebox.showwarning("No AP Selected", "Please select an AP first.")
            return
        
        dialog = tk.Toplevel(self.parent)
        dialog.title("Write Note")
        dialog.geometry("600x500")
        dialog.configure(bg="#FFFFFF")
        dialog.transient(self.parent)
        dialog.grab_set()
        
        main_frame = tk.Frame(dialog, bg="#FFFFFF", padx=20, pady=20)
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        tk.Label(main_frame, text="Write Note", font=("Segoe UI", 14, "bold"),
                bg="#FFFFFF").pack(anchor="w", pady=(0, 15))
        
        # Headline
        tk.Label(main_frame, text="Headline:", font=("Segoe UI", 10, "bold"),
                bg="#FFFFFF").pack(anchor="w", pady=(0, 5))
        
        headline_entry = tk.Entry(main_frame, font=("Segoe UI", 10), relief="flat",
                                 borderwidth=0, highlightthickness=1,
                                 highlightbackground="#CCCCCC",
                                 highlightcolor="#007BFF")
        headline_entry.pack(fill="x", ipady=6)
        headline_entry.focus_set()
        
        # Note content
        tk.Label(main_frame, text="Note:", font=("Segoe UI", 10, "bold"),
                bg="#FFFFFF").pack(anchor="w", pady=(10, 5))
        
        text_widget = scrolledtext.ScrolledText(main_frame, font=("Segoe UI", 10),
                                               wrap=tk.WORD, height=12, relief="flat",
                                               borderwidth=0, highlightthickness=1,
                                               highlightbackground="#CCCCCC",
                                               highlightcolor="#007BFF")
        text_widget.pack(fill="both", expand=True)
        
        # Button frame
        button_frame = tk.Frame(main_frame, bg="#FFFFFF")
        button_frame.pack(fill="x", pady=(10, 0))
        
        def save_note():
            headline = headline_entry.get().strip()
            content = text_widget.get("1.0", tk.END).strip()
            
            if not headline or not content:
                messagebox.showwarning("Missing Data", "Both headline and content are required.",
                                     parent=dialog)
                return
            
            username = self.current_user.get('username') if isinstance(self.current_user, dict) else self.current_user
            success, message, note_id = self.db.add_support_note(self.active_ap, username,
                                                                 headline, content)
            if success:
                self._log(f"Note added: {headline}", "success")
                # Log user activity
                self.db.log_user_activity(username, 'note_create', f'Created note: {headline}', 
                                         ap_id=self.active_ap, success=True)
                # Refresh notes display
                self._load_notes()
                dialog.destroy()
            else:
                messagebox.showerror("Error", f"Failed to save note: {message}", parent=dialog)
        
        # Save button
        tk.Button(button_frame, text="Save Note", command=save_note,
                 bg="#28A745", fg="white", cursor="hand2", padx=20, pady=8,
                 font=("Segoe UI", 10), relief=tk.FLAT, bd=0,
                 activebackground="#218838").pack(side=tk.LEFT, padx=(0, 5))
        
        # Cancel button
        tk.Button(button_frame, text="Cancel", command=dialog.destroy,
                 bg="#6C757D", fg="white", cursor="hand2", padx=20, pady=8,
                 font=("Segoe UI", 10), relief=tk.FLAT, bd=0,
                 activebackground="##5A6268").pack(side=tk.LEFT, padx=5)
    
    def _open_note_window(self, note):
        """Open note details window."""
        # For now, just show a simple dialog with note details
        # TODO: Implement full note window with replies and editing
        self._log(f"Opening note: {note['headline']}")
        
        # Trigger selection callback to show in content panel
        if self.on_selection:
            self.on_selection("note", note)
    
    def _log(self, message, level="info"):
        """Log activity."""
        if self.log_callback:
            self.log_callback("Context Panel", message, level)
